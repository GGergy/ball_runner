import pygame
import random
import os.path
from openpyxl import Workbook, load_workbook
from time import sleep
from timeit import default_timer
if not os.path.isfile('skin.txt'):
    with open('skin.txt', 'w') as file:
        file.write('ball1.png')
if not os.path.isfile('userdata.xlsx'):
    wb = Workbook()
    wb.save('userdata.xlsx')
pygame.init()
screen_width = pygame.display.Info().current_w
screen_height = pygame.display.Info().current_h
screen = pygame.display.set_mode((screen_width, screen_height))
BLACK = (0, 0, 0)
GREEN = (0, 255, 0)
BROWN = (150, 75, 0)
rot = 0.1
menu = True
username = False
clock =  pygame.time.Clock()
wl = {'main2.mp3': (14.3, 30, 'Live Another Day'), 'main3.mp3': (17, 45, 'SiLence'), 'main4.mp3': (11.4, 40, 'Devil Eyes'),
'main5.mp3': (13.5, 30, 'TwiLight'), 'main6.mp3': (6.5, 27, 'DynamIc'), 'main7.mp3': (13, 30, 'Prince of Darkness'), 'main8.mp3': (0, 20, 'OverRide'),
'main9.mp3': (11.5, 18, 'MidNight'), 'main10.mp3': (8.2, 20, 'DisasteR')}
skins = ['ball1.png', 'ball2.png', 'ball3.png', 'ball4.png', 'ball5.png']
with open('skin.txt', 'r') as file:
        ball_skin = file.read()
game_high = 0
FPS = 144
jump_sound = pygame.mixer.Sound('jump.mp3')
count_sound = pygame.mixer.Sound('countdown.mp3')
start_sound = pygame.mixer.Sound('final.mp3')
q_sound = pygame.mixer.Sound('q_tap.mp3')
fail_sound = pygame.mixer.Sound('fail.mp3')
game_user = None
def restart():
    global chunks, emp, game_speed, game_score, game_high, game_hb, in_phonk, bsp, j, did, jr, ball_img, zero_d, sun_color, main_color, colorlist, game_cheats, main_sound, main_song_name, ball_skin, wl, circles, colorlist, game_user, time0
    game_speed = 4 * (60 / FPS)
    game_score = 0
    bsp = 1
    j = False
    jr = False
    did = 0
    game_hb = True
    sun_color = (249, 229, 38)
    game_cheats = False
    in_phonk = False
    main_color = (149, 200, 216)
    colorlist = [GREEN, (173, 255, 47), (52,201,36), (0,69,36), (124,252,0), (0,102,51), (19,136,8), (191,255,0)] 
    if menu:
        pygame.mouse.set_visible(True)
        buttons = [Button(60, 50, 175, 50, 'skin'), Button(screen_width - 200, 50, 100, 50, 'account')]
        main_song_name = False
        def draw():
            pygame.draw.rect(screen, (23,18,48), (0, 0, screen_width, screen_height))
            screen.blit(pygame.image.load(ball_skin), pygame.image.load(ball_skin).get_rect(center=(20, 75)))
            pr_text(txt='select song', n=50, cord=(screen_width // 2, 50))
            pr_text(txt='Select skin', cord=(150, 75), n=30)
            pr_text(txt='Login', cord=(screen_width - 150, 75), n=30)
            if game_user:
                pr_text(txt='Statistic', cord=(screen_width - 150, 200), n=30)
            if game_user:
                pr_text(txt=f'Hi, {game_user.name}', cord=(screen_width - 150, 150), n=30)
            y = 140
            i = 1
            for elem in wl:
                nmb = pr_text(txt=str(i) + '. ' + wl[elem][2], n=30, cord=(screen_width // 2, y), color=(253,149,253), frmt=True)
                buttons.append(Button(nmb.topleft[0] - 5, nmb.topleft[1]- 5, nmb.width + 10, nmb.height + 10, type='song', song_id=i + 1))
                if game_user:
                    buttons.append(Button(screen_width - 210, 175, 120, 50, 'stat'))
                pr_text(txt=str(i) + '. ' + wl[elem][2], n=30, cord=(screen_width // 2, y), color=(253,149,253))
                y += 75
                i += 1
            pr_text(txt='(use mouse)', cord=(screen_width // 2, screen_height - 20), n=25)
            pygame.display.update()
        draw()
        wait_sound = pygame.mixer.Sound('waiting.mp3')
        wait_sound.play()
        while not main_song_name:
            events = pygame.event.get()
            button_invisible = False
            for button in buttons:
                button.is_click()
            for event in events:
                if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                    for button in buttons:
                        if button.is_click():
                            if button.type == 'skin':
                                change_skin()
                                button_invisible = True
                                draw()
                            elif button.type == 'song' and not button_invisible:
                                main_song_name = 'main' + str(button.song_id) + '.mp3'
                                if game_user:
                                    game_user.songs[button.song_id] += 1
                                continue
                            elif button.type == 'account':
                                username = enter_on_keyboard()
                                data = load_workbook('userdata.xlsx')
                                game_user = User(username, data)
                                game_user.get_info()
                                game_high = game_user.high
                                time0 = default_timer()
                                button_invisible = True
                                draw()
                            elif button.type == 'stat':
                                statistic()
                                button_invisible = True
                                draw()
                if event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE or event.type == pygame.QUIT:
                    draw()
                    pygame.draw.rect(screen, (23,18,48), (screen_width // 2 - 175, screen_height // 2 - 150, 350, 200))
                    pygame.draw.rect(screen, (189,208,228), (screen_width // 2 - 175, screen_height // 2 - 150, 350, 200), 7)
                    pygame.mixer.Sound('uhodite.mp3').play()
                    pr_text('exit?', (screen_width // 2, screen_height // 2 - 100), n=50)
                    pr_text('yes', (screen_width // 2 - 100, screen_height // 2), n=40)
                    pr_text('no', (screen_width // 2 + 100, screen_height // 2), n=40)
                    nmb = pr_text('yes', (screen_width // 2 - 100, screen_height // 2), n=40, frmt=True)
                    da = Button(nmb.topleft[0] - 5, nmb.topleft[1]- 5, nmb.width + 10, nmb.height + 10, type='exit')
                    nmb = pr_text('no', (screen_width // 2 + 100, screen_height // 2), n=40, frmt=True)
                    net = Button(nmb.topleft[0] - 5, nmb.topleft[1]- 5, nmb.width + 10, nmb.height + 10, type='exit')
                    pygame.display.update()
                    exit_usl = True
                    while exit_usl:
                        da.is_click()
                        net.is_click()
                        pygame.display.update()
                        events = pygame.event.get()
                        for event in events:
                            if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                                if da.is_click():
                                    pygame.mixer.Sound('pora.mp3').play()
                                    if game_user:
                                        time1 = default_timer()
                                        game_user.time += int(time1 - time0)
                                        game_user.upload_info()
                                    sleep(2)
                                    exit()
                                if net.is_click():
                                    pygame.mixer.Sound('posidim.mp3').play()
                                    exit_usl = False
                                    break
                    button_invisible = True
                    draw()    
        pygame.mouse.set_visible(False)

        pygame.display.update()

        wait_sound.stop()
        count_sound.play()
        pygame.draw.rect(screen, BLACK, (0, 0, screen_width, screen_height))
        pr_text('SPACE to jump', cord=(screen_width // 2, screen_height // 2), n=40)
        pygame.display.update()
        sleep(1)

        count_sound.play()
        pygame.draw.rect(screen, BLACK, (0, 0, screen_width, screen_height))
        pr_text('Q to faster jump', cord=(screen_width // 2, screen_height // 2 - 25), n=40)
        pr_text('E to slower jump', cord=(screen_width // 2, screen_height // 2 + 25), n=40)
        pygame.display.update()
        sleep(1)

        start_sound.play()
        pygame.draw.rect(screen, BLACK, (0, 0, screen_width, screen_height))
        pr_text("LET'S GO", cord=(screen_width // 2, screen_height // 2), n=50)
        pygame.display.update()
        sleep(1.5)

    main_sound = pygame.mixer.Sound(main_song_name)

    chunks = [Chunk(screen_width // 5, screen_height // 2 - 100, GREEN, invin=True)]
    chunk_generated(2)
    emp = Empty(screen_width // 5 + 20, screen_height // 2 - 120, 20, (255, 0, 0)) 
    pygame.draw.rect(screen, BLACK, (0, 0, screen_width, screen_height))
    ball_img = pygame.image.load(ball_skin)
    zero_d = default_timer()
    if game_user:
        game_user.games += 1
    main_sound.play()


def change_skin():
    global ball_skin, game_user
    pygame.draw.rect(screen, (23,18,48), (0, 0, screen_width, screen_height))
    x = screen_width // 2 - len(skins) // 2 * 120
    i = 1
    pr_text(txt='select skin:', n=50, cord=(screen_width // 2, 50))
    skin_buttons = []
    for elem in skins:
        skin_buttons.append(Button(x - 25, 325, 50, 50, song_id=skins.index(elem) + 1))
        screen.blit(pygame.image.load(elem), pygame.image.load(elem).get_rect(center=(x, 300)))
        pr_text(txt=str(i) + '.', cord=(x, 350), color=(253,149,253), n=30)
        x += 100
        i += 1
    pr_text(txt='(use mouse)', cord=(screen_width // 2, screen_height - 30), n=25)
    pygame.display.update()
    ball_skin = False
    while not ball_skin:
        events = pygame.event.get()
        for b in skin_buttons:
            b.is_click()
        pygame.display.update()
        for event in events:
            if event.type == pygame.MOUSEBUTTONDOWN:
                for b in skin_buttons:
                        if b.is_click():
                            ball_skin = 'ball' + str(b.song_id) + '.png'
                            if game_user:
                                game_user.skins[b.song_id] += 1
            elif event.type == pygame.QUIT:
                exit()
    with open('skin.txt', 'w') as file:
                file.write(ball_skin)


def chunk_generated(i=0):
    while chunks[-1].sp_x < screen_width:
        if i > 0:
            chunks.append(Chunk(chunks[-1].sp_x + chunks[-1].ch_width, screen_height // 2 - 100, random.choice(colorlist), invin=True))
        else:
            chunks.append(Chunk(chunks[-1].sp_x + chunks[-1].ch_width, screen_height // 2 - 100, random.choice(colorlist)))
        i -= 1
    for i in range(len(chunks)):
        if chunks[i] and chunks[i].sp_x <= 0 - chunks[i].ch_width:
            chunks[i] = None


def circle_spawn(crcls=screen_width * screen_height // 30000):
    for i in range(crcls):
        if i >= len(circles):
            break
        if circles[i].stade > circles[i].nstade:
            del circles[i]
    for i in range(len(circles), crcls):
        circles.append(Circles())


def draw_background():
    pygame.draw.rect(screen, main_color, (0, 0, screen_width, screen_height))
    pygame.draw.circle(screen, sun_color, (screen_width, 0), 100)
    if in_phonk:
        if default_timer() - zero_d <= main_sound.get_length():
            circle_spawn()
        for circle in circles:
            circle.move()


def enter_on_keyboard(c=(screen_width // 2, screen_height // 2)):
    name = ''
    f = True
    butt = None
    pygame.draw.rect(screen, (23,18,48), (0, 0, screen_width, screen_height))
    pr_text('enter your name:', n=50, cord=(screen_width // 2, 50))
    nmb = pr_text(name, cord=c, n=40, frmt=True)
    butt = Button(nmb.topleft[0] - 5, nmb.topleft[1]- 5, nmb.width + 10, nmb.height + 10, type=None)
    exit_button = Button(60, 50, 75, 50, 'exit')
    pr_text(txt='OK', cord=(100, 75), n=30) 
    pr_text(name, cord=c, n=40)
    butt.light(True)
    pygame.display.update()
    while f or len(name) == 0:
        f = True
        exit_button.is_click()
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                exit()
            elif event.type == pygame.MOUSEBUTTONDOWN:
                if exit_button.is_click():
                    f = False
            elif event.type == pygame.KEYDOWN:
                pygame.draw.rect(screen, (23,18,48), (0, 0, screen_width, screen_height))
                pr_text(txt='OK', cord=(100, 75), n=30) 
                pr_text('enter your name:', n=50, cord=(screen_width // 2, 50))
                if event.key == pygame.K_BACKSPACE:
                    name = name[:-1]
                    nmb = pr_text(name, cord=c, n=40, frmt=True)
                    butt = Button(nmb.topleft[0] - 5, nmb.topleft[1]- 5, nmb.width + 10, nmb.height + 10, type=None)
                    pr_text(name, cord=c, n=40)
                    butt.light(True)
                    continue
                if event.key == pygame.K_SPACE and 0 < len(name):
                        name += ' '
                        continue
                if event.key == pygame.K_RETURN:
                    f = False
                    break
                if len(name) <= 15:
                    try:
                        chr(event.key)
                    except:
                        continue
                    lit = chr(event.key)
                    if lit.isalpha() or lit.isdigit():
                        if pygame.key.get_pressed()[pygame.K_LSHIFT]:
                            name += lit.upper()
                        else:
                            name += lit
            nmb = pr_text(name, cord=c, n=40, frmt=True)
            butt = Button(nmb.topleft[0] - 5, nmb.topleft[1]- 5, nmb.width + 10, nmb.height + 10, type=None)
            pr_text(name, cord=c, n=40)
            butt.light(True)
            pygame.display.update()
    return name


def phonk():
    global in_phonk, colorlist, sun_color, chunks, main_color, circles, game_speed
    game_speed = 10 * (60 / FPS)
    in_phonk = True
    colorlist = [(128,0,128), (89,0,128), (140,0,128), (50,0,128), (113,0,128), (36,0,128), (61,0,128)]
    sun_color = (189,208,228)
    main_color = (23,18,48)
    pygame.draw.rect(screen, BLACK, (0, 0, screen_width, screen_height))
    pr_text(wl[main_song_name][2], n=70, cord=(screen_width // 2, screen_height // 2), color=(255,11,64))
    pygame.display.update()
    chunks.clear()
    chunks = [Chunk(0, screen_height // 2 - 100, (128,0,128), invin=True)]
    circles = []
    emp.njump = 18 * (FPS / 60)
    chunk_generated(4)
    sleep(2)


def pr_text(txt, cord=(320, 240), n=20, color=(255, 255, 255), ctr=True, frmt=False):
    font = pygame.font.SysFont('arial', n)
    text = font.render(txt, True, color)
    if frmt:
        return text.get_rect(center=cord)
    if ctr:
        text_rect = text.get_rect(center=cord)
        screen.blit(text, text_rect)
    else:
        screen.blit(text, cord)


def punch(emp, chunk):
    if chunk.sp and chunk.rect.collidepoint(emp.x, emp.y):
        return game_hb
    return False 


def rotate(screen, image, angle):
    rotated_image = pygame.transform.rotate(image, angle)
    new_rect = rotated_image.get_rect(center=(emp.x, emp.y))
    screen.blit(rotated_image, new_rect)


def statistic():
    def get_key(d, value):
        for k, v in d.items():
            if v == value:
                return k
    key = 'main' + str(get_key(game_user.songs, max(game_user.songs.values()))) + '.mp3'
    pygame.draw.rect(screen, (23,18,48), (0, 0, screen_width, screen_height))
    points = (f'Nickname: {game_user.name}', f'Games played: {game_user.games}', f'Time in game: {game_user.time // 3600}h {game_user.time % 3600 // 60}m {game_user.time % 60}s',
    f'Favourite song: {wl[key][2]}', 'Favourite skin:')
    x = 150
    y = 300
    pr_text('Your Statistics:', cord=(screen_width // 2, 100), n=50)
    for i in range(len(points)):
        pr_text(points[i], (x, y), n=35, ctr=False)
        y += 50
    key = 'ball' + str(get_key(game_user.skins, max(game_user.skins.values()))) + '.png'
    img = pygame.image.load(key)
    screen.blit(img, img.get_rect(center=(400, 520)))
    pygame.display.update()
    f = True
    while f:
        for e in pygame.event.get():
            if e.type == pygame.KEYDOWN and e.key == pygame.K_ESCAPE:
                f = False

class Empty:
    def __init__(self, x, y, rad, col) -> None:
        self.x = x
        self.y = y
        self.rad = rad
        self.col = col
        self.njump = 45 * (FPS / 60)
        self.jumpstade = 1
        self.last = game_speed
        self.futurejump = False

    def jump_reset(self, upd):
        if self.njump > 10 * (FPS / 60):
            self.njump -= 2 * upd * (FPS / 60)
        if upd < 0:
            self.njump -= 2 * upd * (FPS / 60)

    def jump(self):
        if self.jumpstade == 1:
            jump_sound.play()
        if self.jumpstade <= ((self.njump - (5 * (FPS / 60))) / 2):
            self.y -= 85 / ((self.njump - (5 * (FPS / 60))) / 2)
        elif self.jumpstade > ((self.njump - (5 * (FPS / 60))) / 2) + 5 * (FPS / 60):
            self.y += 85 / ((self.njump - (5 * (FPS / 60))) / 2)  

        if self.jumpstade >= self.njump:
            self.jumpstade = 1
            self.y = screen_height // 2 - 120
        else:
            self.jumpstade += 1
        
    def is_score(self, chunk):
        global game_score, did
        if chunk and chunk.sp_x <= self.x <= chunk.sp_x + chunk.ch_width and chunk.sp:
            if chunk.block_spawn + chunk.sp_x - self.rad <= self.x <= chunk.block_spawn + 20 + chunk.sp_x + self.rad and chunks.index(chunk) != did:
                game_score += 1
                did = chunks.index(chunk)


class Chunk:
    def __init__(self, sp_x, sp_y, c_col, invin=False):
        global bsp
        self.sp_x = sp_x
        self.sp_y = sp_y
        self.c_col = c_col
        self.ch_width = 175
        self.ch_height = 70
        if invin:
            self.sp = False
        else:
            self.sp = bsp % 2 == 0
        bsp += 1

        pygame.draw.rect(screen, self.c_col, (sp_x, sp_y, self.ch_width, self.ch_height))
        if self.sp:
            self.block_spawn = random.randint(0, self.ch_width - 20)
            self.img = pygame.image.load('block_texture.png')
            self.rect = self.img.get_rect(center = (self.sp_x + self.block_spawn + 10, self.sp_y - 30))        

    def chunk_move(self):
        if self.sp_x + self.ch_width >=0:
            self.sp_x -= game_speed
            pygame.draw.rect(screen, self.c_col, (self.sp_x, self.sp_y, self.ch_width, self.ch_height))
            if self.sp:
                self.rect = self.img.get_rect(center = (self.sp_x + self.block_spawn + 10, self.sp_y - 30))
                screen.blit(self.img, self.rect)
    

class Circles:
    def __init__(self):
        self.rad = 0
        self.maxrad = 75
        self.nstade = int(wl[main_song_name][1] * (FPS / 60))
        self.stade = 1
        self.y = random.randint(0, screen_height)
        self.x = random.randint(0, screen_width)
        self.color = random.choice(colorlist)
        pygame.draw.circle(screen, self.color, (self.x, self.y), self.rad)

    def move(self):
        if self.stade <= (self.nstade - 5) // 2:
            self.rad += self.maxrad / ((self.nstade - 5) // 2)
        elif self.stade >= (self.nstade - 5) // 2 + 6:
            self.rad -= self.maxrad / ((self.nstade - 5) // 2)
        self.stade += 1
        pygame.draw.circle(screen, self.color, (self.x, self.y), self.rad)

class Button:
    def __init__(self, x, y, w ,h, type=False, song_id=False) -> None:
        self.x = x
        self.y = y
        self.w = w
        self.h = h
        self.song_id = song_id
        self.type = type
        self.heatbox = pygame.Rect(x, y, w, h)
        self.last_light = False

    
    def is_click(self):
        x, y = pygame.mouse.get_pos()
        if self.heatbox.collidepoint(x, y):
            self.light(True)
            return True
        else:
            self.light(False)

    def light(self, show):
        color = (168,162,252) if show else (23,18,48)
        pygame.draw.rect(screen, color, (self.x, self.y, self.w, self.h), 5)
        if show != self.last_light:
            pygame.display.update()
            self.last_light = show


class User:
    def __init__(self, name, doc) -> None:
        self.name = name
        self.doc = doc
    
    def get_info(self):
        self.high = 0
        self.games = 0
        self.time = 0
        self.songs = {}
        self.skins = {}
        if self.name not in self.doc.sheetnames:
            self.doc.create_sheet(self.name)
            sheet = self.doc[self.name]
            for i in range(1, 4):
                sheet['A' + str(i)].value = 0
            for i in range(1, len(wl) + 1):
                sheet['B' + str(i)].value = 0
                self.songs[i + 1] = 0
            for i in range(1, len(skins) + 1):
                sheet['C' + str(i)].value = 0
                self.skins[i] = 0
        else:
            sheet = self.doc[self.name]
            self.high = sheet['A1'].value
            self.games = sheet['A2'].value
            self.time = sheet['A3'].value
            for i in range(1, len(wl) + 1):
                self.songs[i + 1] = sheet['B' + str(i)].value
            for i in range(1, len(skins) + 1):
                self.skins[i] = sheet['C' + str(i)].value

    def upload_info(self):
        sheet = self.doc[self.name]
        sheet['A1'].value = self.high
        sheet['A2'].value = self.games
        sheet['A3'].value = self.time
        for i in range(1, len(wl) + 1):
            sheet['B' + str(i)].value = self.songs[i + 1] = 0
        for i in range(1, len(skins) + 1):
            sheet['C' + str(i)].value = self.skins[i] = 0
        self.doc.save('userdata.xlsx')

restart()
while True: 
    #print(default_timer() - zero_d)
    if default_timer() - zero_d >= wl[main_song_name][0] and not in_phonk:
        phonk() 

    for i in range(len(chunks) - 1, -1, -1):
        if not chunks[i]:
            break 

        if punch(emp, chunks[i]):
            main_sound.stop()
            fail_sound.play()

            if i == did:
                game_score -= 1
            if game_score > game_high and not game_cheats:
                game_high = game_score
                if game_user:
                    game_user.high = game_high

            pr_text(f'game over, score: {game_score}  hight: {game_high}', n=36, cord=(screen_width // 2, 130))
            pr_text('CTRL to restart, ESC to exit in main menu', n=30, cord=(screen_width // 2, 200))

            pygame.display.update()
            events = pygame.event.get()
            f = False
            for e in events:
                if e.type == pygame.KEYDOWN and e.key == pygame.K_LCTRL:
                    f = True
            while not f:
                events = pygame.event.get()
                for e in events:
                    if e.type == pygame.KEYDOWN:
                        if e.key == pygame.K_LCTRL:
                            f = True
                            menu = False
                        if e.key == pygame.K_ESCAPE:
                            f = True
                            menu = True
                    if e.type == pygame.QUIT:
                        exit()
                pass
            fail_sound.stop()
            restart()
            break


    events = pygame.event.get()
    for event in events:
        if event.type == pygame.QUIT:
            exit()
        if event.type == pygame.KEYDOWN:
            if event.key in (pygame.K_SPACE, pygame.K_UP) and not j:
                j = True
            if event.key in (pygame.K_q, pygame.K_e):
                jr = True
                if event.key == pygame.K_q:
                    upd = 1
                else:
                    upd = -1
            if event.key == pygame.K_h: 
                game_hb = not game_hb
                game_cheats = True
            if event.key == pygame.K_ESCAPE:
                start = default_timer()
                pygame.mixer.pause()
                pr_text(txt=('game paused...'), n=75, cord=(screen_width // 2, screen_height // 2 - 100))
                if random.randint(1, 10) == 1:
                    pr_text(txt='поздравляю, ты увидел пасхалку)', n=30, cord=(screen_width // 2, screen_height - 100))
                pygame.display.update()
                usl = True
                while usl:
                    for event in pygame.event.get():
                        if event.type == pygame.QUIT:
                            exit()
                        if event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE:
                            usl = False
                            break
                pygame.mixer.unpause()
                stop = default_timer()
                zero_d += stop - start            
        if event.type == pygame.MOUSEBUTTONDOWN:
            if j and emp.jumpstade >= emp.njump * 0.75:
                emp.futurejump = True
            else:
                j = True

    if emp.futurejump and not j:
        j = True
        emp.futurejump = False

    if j:
        emp.jump()
        if emp.jumpstade == 1:
            j = False
    
    for c in chunks:
        emp.is_score(c)
         
    if jr and emp.jumpstade == 1:
        q_sound.play()
        emp.jump_reset(upd)
        jr = False

    draw_background()
    for ch in chunks:
        if ch:
            ch.chunk_move()
    chunk_generated()

    rotate(screen, ball_img, -rot % 360)
    if not j:
        rot += 5 * (60 / FPS)
    else:
        rot += 2 * (60 / FPS)
    if default_timer() - zero_d >= main_sound.get_length() // 1 + 2:
        if game_score > game_high and not game_cheats:
            game_high = game_score
            if game_user:
                game_user.high = game_high
        pr_text(txt=f'complete! score: {game_score}', n=70, cord=(screen_width // 2, screen_height // 2 - 100), color=(1,112,34))
        menu = True
        pygame.display.update()
        sleep(3)
        restart()
        pygame.display.update()
        continue
    txt = f'this-{game_score}' + (f'high-{game_high}' if game_user else '  log in to see statistic')
    pr_text(txt, cord=(50, 50), ctr=False)
    if game_user:
        pr_text(f'playing - {game_user.name}', cord=(screen_width - 250, 50), ctr=False)
    if not game_hb:
        pr_text('!GOD MODE!', cord=(screen_width // 2, 50), n=40, color=(227,20,20))
    pygame.display.update()
    clock.tick(FPS) 
